"""Raw file readers.

Bronze preserves source fidelity, so these readers do as little as possible: no dtype
inference, no NA coercion, no header cleanup. A cell comes back as whatever the file
said it was — a string, a number, or a ``datetime`` when the workbook itself carried a
date format. Interpretation happens downstream, where it can be logged.

Real submissions do not start at cell A1. A title row, a blank spacer, and a merged
banner above the header are all routine, so :func:`read_tables` locates the header row
by scoring the first few rows against the alias pool rather than assuming row 0.
"""

from __future__ import annotations

import csv
import math
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

from sbfp_platform.utils.text import normalize_header

#: How far down a sheet to look for the header row before giving up and using the first
#: non-empty row.
HEADER_SCAN_ROWS = 10

#: Encodings tried in order when a CSV is not valid UTF-8. Field submissions routinely
#: arrive as Windows-1252 out of Excel.
CSV_ENCODINGS = ("utf-8-sig", "cp1252", "latin-1")

SUPPORTED_SUFFIXES = {".xlsx": "xlsx", ".csv": "csv"}


@dataclass(frozen=True)
class RawTable:
    """One sheet (or one CSV) as read off disk, before any mapping.

    Attributes:
        sheet_name: Worksheet name, or ``None`` for a CSV.
        headers: Column headers exactly as written, de-duplicated so they can key a
            JSON payload. Blank headers become ``column_<n>``.
        rows: Data rows, each padded or trimmed to ``len(headers)``. The index of a row
            in this list is its ``source_row_number``.
        header_row_index: 0-based index of the row the headers came from, kept so a
            bronze row can be traced to a spreadsheet cell.
    """

    sheet_name: str | None
    headers: list[str]
    rows: list[list[object]] = field(default_factory=list)
    header_row_index: int = 0


def file_type_of(path: Path) -> str | None:
    """Return ``"xlsx"``/``"csv"``, or ``None`` if the platform does not read this file."""
    return SUPPORTED_SUFFIXES.get(path.suffix.lower())


def cell_to_text(cell: object) -> str:
    """Render a cell as the text a person would have seen in the spreadsheet.

    Integral floats lose their ``.0``: openpyxl reports a whole-number cell as
    ``43262.0``, but the school typed ``43262``, and bronze stores what was typed.
    """
    if cell is None:
        return ""
    if isinstance(cell, float):
        if math.isnan(cell) or math.isinf(cell):
            return ""
        if cell.is_integer():
            return str(int(cell))
    return str(cell).strip()


def read_tables(path: Path, header_keys: set[str]) -> list[RawTable]:
    """Read every populated sheet in ``path``.

    Args:
        path: File to read. Must be ``.xlsx`` or ``.csv``.
        header_keys: Folded alias keys (see :func:`~sbfp_platform.utils.text.normalize_header`)
            used to score candidate header rows. Pass the union across all datasets:
            header detection must not presuppose the classification it feeds.

    Returns:
        One :class:`RawTable` per non-empty sheet, in workbook order. A CSV yields at
        most one.

    Raises:
        ValueError: If the suffix is not supported.
    """
    kind = file_type_of(path)
    if kind == "xlsx":
        return _read_xlsx(path, header_keys)
    if kind == "csv":
        return _read_csv(path, header_keys)
    raise ValueError(f"Unsupported file type for ingestion: {path.name}")


# --------------------------------------------------------------------------------------
# Format-specific readers
# --------------------------------------------------------------------------------------


def _read_xlsx(path: Path, header_keys: set[str]) -> list[RawTable]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        tables = []
        for sheet in workbook.worksheets:
            grid = [list(row) for row in sheet.iter_rows(values_only=True)]
            table = _build_table(grid, header_keys, sheet_name=sheet.title)
            if table is not None:
                tables.append(table)
        return tables
    finally:
        workbook.close()


def _read_csv(path: Path, header_keys: set[str]) -> list[RawTable]:
    text = _read_text(path)
    grid: list[list[object]] = [list(row) for row in csv.reader(text.splitlines())]
    table = _build_table(grid, header_keys, sheet_name=None)
    return [table] if table is not None else []


def _read_text(path: Path) -> str:
    last_error: UnicodeDecodeError | None = None
    for encoding in CSV_ENCODINGS:
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError as exc:  # pragma: no cover - depends on file bytes
            last_error = exc
    raise last_error  # pragma: no cover


# --------------------------------------------------------------------------------------
# Header location and row shaping
# --------------------------------------------------------------------------------------


def _build_table(
    grid: list[list[object]], header_keys: set[str], *, sheet_name: str | None
) -> RawTable | None:
    header_index = _locate_header_row(grid, header_keys)
    if header_index is None:
        return None

    headers = _dedupe_headers(grid[header_index])
    width = len(headers)
    rows = [_fit(row, width) for row in grid[header_index + 1 :]]

    # Trailing blank rows are an artifact of how spreadsheets store used ranges, not
    # data. Interior blanks are kept so source_row_number stays aligned with the sheet.
    while rows and _is_blank_row(rows[-1]):
        rows.pop()

    return RawTable(
        sheet_name=sheet_name, headers=headers, rows=rows, header_row_index=header_index
    )


def _locate_header_row(grid: list[list[object]], header_keys: set[str]) -> int | None:
    """Pick the header row: the best-scoring row in the scan window, else the first."""
    best_index: int | None = None
    best_score = 0
    first_populated: int | None = None

    for index, row in enumerate(grid[:HEADER_SCAN_ROWS]):
        if _is_blank_row(row):
            continue
        if first_populated is None:
            first_populated = index
        score = sum(
            1
            for cell in row
            if cell_to_text(cell) and normalize_header(cell_to_text(cell)) in header_keys
        )
        if score > best_score:
            best_index, best_score = index, score

    if best_index is not None:
        return best_index
    return first_populated


def _dedupe_headers(row: list[object]) -> list[str]:
    """Name every column uniquely so headers can key a JSON payload."""
    seen: dict[str, int] = {}
    headers: list[str] = []
    for position, cell in enumerate(row, start=1):
        name = cell_to_text(cell) or f"column_{position}"
        count = seen.get(name, 0) + 1
        seen[name] = count
        headers.append(name if count == 1 else f"{name}__{count}")
    return headers


def _fit(row: list[object], width: int) -> list[object]:
    if len(row) < width:
        return [*row, *([None] * (width - len(row)))]
    return row[:width]


def _is_blank_row(row: list[object]) -> bool:
    return all(cell_to_text(cell) == "" for cell in row)
